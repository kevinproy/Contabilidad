"""
Sistema Contable Web (Fase 1 - Flask con PostgreSQL)
"""

from __future__ import annotations

import os
from typing import Any, Dict, List, Optional, Callable

from flask import Flask, render_template, session, redirect, url_for, flash, request, send_file
from functools import wraps

from services.db import ensure_db_schema, db_connect
from blueprints.estado_cuenta import bp as estado_bp
from blueprints.auth import bp as auth_bp, init_auth
from io import BytesIO
import json
from datetime import datetime, timedelta
try:
    import pandas as pd
except Exception:
    pass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ==========================
# Configuración/Constantes
# ==========================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Columnas de salida estándar en vistas/exportación
COLUMNAS_SALIDA = [
    "FECHA",
    "DOCTO",
    "DETALLE",
    "DEBE",
    "HABER",
    "SALDO",
    "INT. AG.",
    "DIM",
    "CONDICION DE PAGO",
]


def create_app() -> Flask:
    app = Flask(__name__)
    app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")

    @app.get("/__db_check")
    def __db_check():
        try:
            host = os.environ.get("PGHOST")
            port = os.environ.get("PGPORT")
            dbn = os.environ.get("PGDATABASE")
            usr = os.environ.get("PGUSER")
            with db_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT version()")
                    ver = cur.fetchone()[0]
            return {"ok": True, "host": host, "port": port, "database": dbn, "user": usr, "version": ver}
        except Exception as exc:
            return {"ok": False, "error": str(exc)}, 500

    try:
        ensure_db_schema()
    except Exception as exc:
        print(f"[WARN] No se pudo asegurar el esquema en DB: {exc}")

    # ==========================
    # Utilidades DB (helpers)
    # ==========================

    def db_get_or_create_cliente(nombre: str) -> Optional[int]:
        nombre_norm = (nombre or "").strip()
        if not nombre_norm:
            return None
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id_cliente FROM clientes WHERE nombre_completo=%s", (nombre_norm,))
                row = cur.fetchone()
                if row:
                    return int(row["id_cliente"]) if isinstance(row, dict) else int(row[0])
                cur.execute(
                    "INSERT INTO clientes (nombre_completo) VALUES (%s) ON CONFLICT (nombre_completo) DO NOTHING RETURNING id_cliente",
                    (nombre_norm,),
                )
                r2 = cur.fetchone()
                if r2:
                    conn.commit()
                    return int(r2["id_cliente"]) if isinstance(r2, dict) else int(r2[0])
                # Si hubo conflicto, volver a consultar
                cur.execute("SELECT id_cliente FROM clientes WHERE nombre_completo=%s", (nombre_norm,))
                row = cur.fetchone()
                conn.commit()
                return int(row["id_cliente"]) if row else None

    def db_insert_movimiento(
        id_cliente: int,
        fecha: Optional[str],
        tipo: str,
        monto: float,
        descripcion: Optional[str],
        docto: Optional[str],
        int_ag: Optional[str],
        dim: Optional[str],
        condicion_de_pago: Optional[str],
    ) -> bool:
        """Inserta movimiento idempotente. Retorna True si insertó, False si existía."""
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    (
                        "INSERT INTO movimientos (id_cliente, fecha, tipo_de_movimiento, monto, descripcion, docto, int_ag, dim, condicion_de_pago) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                        "ON CONFLICT (id_cliente, fecha, tipo_de_movimiento, monto, descripcion) DO NOTHING RETURNING id_movimiento"
                    ),
                    (id_cliente, fecha, tipo, float(monto or 0), descripcion, docto, int_ag, dim, condicion_de_pago),
                )
                r = cur.fetchone()
                conn.commit()
                return bool(r)

    # Filtro Jinja para formateo numérico estilo ES (miles con punto, decimales con coma)
    def format_number_es(value: Any) -> str:
        try:
            num = float(value)
        except Exception:
            return str(value)
        text = f"{num:,.2f}"
        return text.replace(",", "X").replace(".", ",").replace("X", ".")

    app.jinja_env.filters["es_num"] = format_number_es

    # Permisos del sistema
    PERM_ESTADO_VIEW = "estado:view"
    PERM_ESTADO_UPLOAD = "estado:upload"
    PERM_ESTADO_EXPORT = "estado:export"
    PERM_ESTADO_ANULADAS_VIEW = "estado:anuladas:view"
    PERM_ESTADO_SALDOS_MANAGE = "estado:saldos:manage"
    PERM_ESTADO_PREFS_MANAGE = "estado:prefs:manage"
    PERM_PLANILLA_VIEW = "planilla:view"
    PERM_COMISIONES_VIEW = "comisiones:view"
    PERM_DIM_VIEW = "dim:view"
    PERM_DASHBOARD_VIEW = "dashboard:view"
    PERM_PLANILLA_EDIT = "planilla:edit"
    PERM_PLANILLA_EMPL_VIEW = "planilla:empleados:view"
    PERM_PLANILLA_PERIODO_VIEW = "planilla:periodo:view"

    ROLE_PERMISSIONS: Dict[str, List[str]] = {
        # Tabla de referencia por si se usa roles estáticos en futuro
        "superadmin": [
            PERM_DASHBOARD_VIEW,
            PERM_ESTADO_VIEW,
            PERM_ESTADO_UPLOAD,
            PERM_PLANILLA_VIEW,
            PERM_COMISIONES_VIEW,
            PERM_DIM_VIEW,
        ],
        "consultor": [PERM_DASHBOARD_VIEW, PERM_ESTADO_VIEW],
        "operador_estado": [PERM_DASHBOARD_VIEW, PERM_ESTADO_VIEW, PERM_ESTADO_UPLOAD],
    }

    def usuario_actual() -> Optional[Dict[str, Any]]:
        return session.get("user")

    def usuario_tiene_permiso(user: Dict[str, Any], permiso: str) -> bool:
        if not user:
            return False
        # Maestro siempre puede
        if user.get("is_master"):
            return True
        # Leer del listado de permisos en sesión
        perms = user.get("perms") or []
        return permiso in perms

    def login_required(view: Callable) -> Callable:
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not usuario_actual():
                flash("Inicie sesión para continuar.", "error")
                return redirect(url_for("auth.login"))
            return view(*args, **kwargs)
        return wrapped

    def permission_required(permiso: str) -> Callable:
        def decorator(view: Callable) -> Callable:
            @wraps(view)
            def wrapped(*args, **kwargs):
                user = usuario_actual()
                if not user:
                    flash("Inicie sesión para continuar.", "error")
                    return redirect(url_for("auth.login"))
                if not usuario_tiene_permiso(user, permiso):
                    flash("No tiene permisos para acceder a esta sección.", "error")
                    return redirect(url_for("index"))
                return view(*args, **kwargs)
            return wrapped
        return decorator

    @app.context_processor
    def inject_context():
        user = usuario_actual()
        def can(p: str) -> bool:
            return usuario_tiene_permiso(user, p) if user else False
        return dict(current_user=user, can=can,
                    PERM_ESTADO_VIEW=PERM_ESTADO_VIEW,
                    PERM_ESTADO_UPLOAD=PERM_ESTADO_UPLOAD,
                    PERM_ESTADO_EXPORT=PERM_ESTADO_EXPORT,
                    PERM_ESTADO_ANULADAS_VIEW=PERM_ESTADO_ANULADAS_VIEW,
                    PERM_ESTADO_SALDOS_MANAGE=PERM_ESTADO_SALDOS_MANAGE,
                    PERM_ESTADO_PREFS_MANAGE=PERM_ESTADO_PREFS_MANAGE,
                    PERM_PLANILLA_VIEW=PERM_PLANILLA_VIEW,
                    PERM_PLANILLA_EDIT=PERM_PLANILLA_EDIT,
                    PERM_PLANILLA_EMPL_VIEW=PERM_PLANILLA_EMPL_VIEW,
                    PERM_PLANILLA_PERIODO_VIEW=PERM_PLANILLA_PERIODO_VIEW,
                    PERM_COMISIONES_VIEW=PERM_COMISIONES_VIEW,
                    PERM_DIM_VIEW=PERM_DIM_VIEW,
                    PERM_DASHBOARD_VIEW=PERM_DASHBOARD_VIEW)

    init_auth(app)

    # ==========================
    # Utilidades de datos
    # ==========================

    def cargar_json_estado_de_cuenta() -> List[Dict[str, Any]]:
        # Fuente única: base de datos
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT m.id_movimiento AS "ID",
                           c.nombre_completo AS "CLIENTE",
                           COALESCE(TO_CHAR(m.fecha, 'YYYY-MM-DD'), '') AS "FECHA",
                           m.docto AS "DOCTO",
                           m.descripcion AS "DETALLE",
                           m.int_ag AS "INT. AG.",
                           m.dim AS "DIM",
                           COALESCE(TO_CHAR(m.condicion_de_pago, 'YYYY-MM-DD'), '') AS "CONDICION DE PAGO",
                           m.tipo_de_movimiento,
                           m.monto,
                           m.order_index AS "ORDER",
                           COALESCE(m.mark_debe,0) AS MARK_DEBE,
                           COALESCE(m.mark_haber,0) AS MARK_HABER,
                           COALESCE(m.mark_saldo,0) AS MARK_SALDO
                    FROM movimientos m
                    JOIN clientes c ON c.id_cliente = m.id_cliente
                    WHERE m.anulada_en IS NULL
                    """
                )
                rows = cur.fetchall()
                out: List[Dict[str, Any]] = []
                for r in rows:
                    debe = float(r["monto"]) if r["tipo_de_movimiento"] == "CARGO" else 0.0
                    haber = float(r["monto"]) if r["tipo_de_movimiento"] == "PAGO" else 0.0
                    out.append({
                        "ID": str(r["ID"]),
                        "CLIENTE": r["CLIENTE"],
                        "FECHA": r["FECHA"],
                        "DOCTO": r["DOCTO"] or "",
                        "DETALLE": r["DETALLE"] or "",
                        "INT. AG.": r["INT. AG."] or "",
                        "DIM": r["DIM"] or "",
                        "CONDICION DE PAGO": r["CONDICION DE PAGO"] or "",
                        "DEBE": round(debe, 2),
                        "HABER": round(haber, 2),
                        "SALDO": 0.0,
                        "ORDER": r.get("ORDER"),
                        "MARK_DEBE": int(r.get("MARK_DEBE", 0)),
                        "MARK_HABER": int(r.get("MARK_HABER", 0)),
                        "MARK_SALDO": int(r.get("MARK_SALDO", 0)),
                    })
                return out

    # Anuladas
    def cargar_json_anuladas() -> List[Dict[str, Any]]:
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT c.nombre_completo AS "CLIENTE",
                           COALESCE(TO_CHAR(m.fecha,'YYYY-MM-DD'),'') AS "FECHA",
                           m.docto AS "DOCTO",
                           m.descripcion AS "DETALLE",
                           m.int_ag AS "INT. AG.",
                           m.dim AS "DIM",
                           m.tipo_de_movimiento,
                           m.monto,
                           COALESCE(TO_CHAR(m.anulada_en,'YYYY-MM-DD HH24:MI:SS'),'') AS "ANULADO_EN"
                    FROM movimientos m
                    JOIN clientes c ON c.id_cliente = m.id_cliente
                    WHERE m.anulada_en IS NOT NULL
                    ORDER BY m.anulada_en DESC
                    """
                )
                rows = cur.fetchall()
                out: List[Dict[str, Any]] = []
                for r in rows:
                    debe = float(r["monto"]) if r["tipo_de_movimiento"] == "CARGO" else 0.0
                    haber = float(r["monto"]) if r["tipo_de_movimiento"] == "PAGO" else 0.0
                    out.append({
                        "FECHA": r["FECHA"],
                        "DOCTO": r["DOCTO"] or "",
                        "DETALLE": r["DETALLE"] or "",
                        "INT. AG.": r["INT. AG."] or "",
                        "DIM": r["DIM"] or "",
                        "DEBE": round(debe, 2),
                        "HABER": round(haber, 2),
                        "SALDO": 0.0,
                        "ANULADO_EN": r["ANULADO_EN"],
                    })
                return out

    # Saldos iniciales por cliente {cliente: monto}
    def cargar_saldos_iniciales() -> Dict[str, Dict[str, Any]]:
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT c.nombre_completo AS cliente, s.monto, s.lado,
                           COALESCE(TO_CHAR(s.fecha,'YYYY-MM-DD'),'2025-01-01') AS fecha
                    FROM saldos_iniciales s
                    JOIN clientes c ON c.id_cliente = s.id_cliente
                    """
                )
                rows = cur.fetchall()
                return {r["cliente"]: {"monto": float(r["monto"] or 0), "lado": r["lado"], "fecha": r["fecha"]} for r in rows}

    def guardar_saldos_iniciales(mapa: Dict[str, Dict[str, Any]]) -> None:
        with db_connect() as conn:
            with conn.cursor() as cur:
                for cliente, info in mapa.items():
                    cur.execute("SELECT id_cliente FROM clientes WHERE nombre_completo=%s", (cliente,))
                    row = cur.fetchone()
                    if not row:
                        cur.execute("INSERT INTO clientes (nombre_completo) VALUES (%s) RETURNING id_cliente", (cliente,))
                        cid = cur.fetchone()["id_cliente"]
                    else:
                        cid = row["id_cliente"]
                    cur.execute(
                        """
                        INSERT INTO saldos_iniciales (id_cliente, monto, lado, fecha)
                        VALUES (%s,%s,%s,%s)
                        ON CONFLICT (id_cliente) DO UPDATE SET monto=EXCLUDED.monto, lado=EXCLUDED.lado, fecha=EXCLUDED.fecha
                        """,
                        (cid, float(info.get("monto", 0) or 0), str(info.get("lado","haber")), str(info.get("fecha","2025-01-01"))),
                    )
            conn.commit()

    # Helpers de Excel y allowed_file: ahora residen en services/estado_service.py y blueprint

    # ==========================
    # Rutas
    # ==========================

    @app.get("/")
    @login_required
    def index():
        # Home: si tiene dashboard, ir a dashboard; si no, planilla o estado según permisos
        user = usuario_actual()
        if usuario_tiene_permiso(user, PERM_DASHBOARD_VIEW):
            return redirect(url_for("dashboard"))
        if usuario_tiene_permiso(user, PERM_PLANILLA_VIEW):
            now = datetime.now()
            periodo = now.strftime("%Y%m")
            return redirect(url_for("planilla_periodo", periodo=periodo))
        if usuario_tiene_permiso(user, PERM_ESTADO_VIEW):
            return redirect(url_for("estado.ver_tabla_estado_cuenta"))
        flash("No tiene permisos asignados. Contacte al administrador.", "error")
        return redirect(url_for("auth.logout"))

    # Registrar blueprints
    app.register_blueprint(auth_bp)
    app.register_blueprint(estado_bp)

    # ===============
    # Admin de usuarios
    # ===============
    @app.get("/admin/users")
    @login_required
    def admin_users():
        user = usuario_actual()
        if not (user and user.get("is_master")):
            flash("Solo el administrador maestro puede gestionar usuarios.", "error")
            return redirect(url_for("index"))
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id_user, username, is_master, is_active FROM users ORDER BY id_user ASC")
                users = cur.fetchall()
                cur.execute("SELECT code, description FROM permissions ORDER BY code ASC")
                perms = cur.fetchall()
                cur.execute("SELECT id_user, perm_code FROM user_permissions")
                rows = cur.fetchall()
        grants = {}
        for r in rows:
            uid = r.get("id_user") if isinstance(r, dict) else r[0]
            code = r.get("perm_code") if isinstance(r, dict) else r[1]
            grants.setdefault(uid, set()).add(code)
        return render_template("admin_users.html", users=users, perms=perms, grants=grants)

    @app.post("/admin/users/create")
    @login_required
    def admin_users_create():
        user = usuario_actual()
        if not (user and user.get("is_master")):
            return {"ok": False, "error": "Solo maestro"}, 403
        from werkzeug.security import generate_password_hash
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if not username or not password:
            flash("Usuario y contraseña requeridos.", "error")
            return redirect(url_for("admin_users"))
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO users (username, password_hash, is_master) VALUES (%s,%s,FALSE) ON CONFLICT (username) DO NOTHING RETURNING id_user",
                    (username, generate_password_hash(password)),
                )
                row = cur.fetchone()
                conn.commit()
        if not row:
            flash("El usuario ya existe.", "error")
        else:
            flash("Usuario creado.", "success")
        return redirect(url_for("admin_users"))

    @app.post("/admin/users/perms")
    @login_required
    def admin_users_perms():
        user = usuario_actual()
        if not (user and user.get("is_master")):
            return {"ok": False, "error": "Solo maestro"}, 403
        uid = request.form.get("id_user")
        granted: List[str] = request.form.getlist("perms")
        try:
            uid_i = int(uid)
        except Exception:
            flash("Usuario inválido.", "error")
            return redirect(url_for("admin_users"))
        # No permitir cambiar permisos del maestro
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT is_master FROM users WHERE id_user=%s", (uid_i,))
                rowm = cur.fetchone()
                if rowm and (rowm.get("is_master") if isinstance(rowm, dict) else rowm[0]):
                    flash("No se pueden modificar permisos del usuario maestro.", "error")
                    return redirect(url_for("admin_users"))
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM user_permissions WHERE id_user=%s", (uid_i,))
                for code in granted:
                    cur.execute(
                        "INSERT INTO user_permissions (id_user, perm_code) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                        (uid_i, code),
                    )
                conn.commit()
        flash("Permisos actualizados.", "success")
        return redirect(url_for("admin_users"))

    @app.post("/admin/users/update")
    @login_required
    def admin_users_update():
        user = usuario_actual()
        if not (user and user.get("is_master")):
            return {"ok": False, "error": "Solo maestro"}, 403
        uid = request.form.get("id_user")
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        try:
            uid_i = int(uid)
        except Exception:
            flash("Usuario inválido.", "error")
            return redirect(url_for("admin_users"))
        from werkzeug.security import generate_password_hash
        with db_connect() as conn:
            with conn.cursor() as cur:
                if username:
                    cur.execute("UPDATE users SET username=%s WHERE id_user=%s", (username, uid_i))
                if password:
                    cur.execute("UPDATE users SET password_hash=%s WHERE id_user=%s", (generate_password_hash(password), uid_i))
                conn.commit()
        flash("Usuario actualizado.", "success")
        return redirect(url_for("admin_users"))

    @app.post("/admin/users/delete")
    @login_required
    def admin_users_delete():
        user = usuario_actual()
        if not (user and user.get("is_master")):
            return {"ok": False, "error": "Solo maestro"}, 403
        uid = request.form.get("id_user")
        try:
            uid_i = int(uid)
        except Exception:
            flash("Usuario inválido.", "error")
            return redirect(url_for("admin_users"))
        # No permitir borrar al propio usuario ni a un maestro
        if user.get("id") == uid_i:
            flash("No puede eliminar su propia cuenta.", "error")
            return redirect(url_for("admin_users"))
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT is_master FROM users WHERE id_user=%s", (uid_i,))
                rowm = cur.fetchone()
                if rowm and (rowm.get("is_master") if isinstance(rowm, dict) else rowm[0]):
                    flash("No se puede eliminar al usuario maestro.", "error")
                    return redirect(url_for("admin_users"))
                cur.execute("DELETE FROM users WHERE id_user=%s", (uid_i,))
                conn.commit()
        flash("Usuario eliminado.", "success")
        return redirect(url_for("admin_users"))

    # La carga de Excel se maneja en el blueprint `estado`

    # La tabla principal se sirve desde el blueprint `estado`

    @app.post("/estado-cuenta/reordenar/<cliente>")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def reordenar_cliente(cliente: str):
        payload = request.get_json(silent=True) or {}
        ids = payload.get("ids", [])
        if not isinstance(ids, list) or not ids:
            return {"ok": False, "error": "Lista de IDs vacía"}, 400
        # Asignar order_index secuencial según posición en lista
        with db_connect() as conn:
            with conn.cursor() as cur:
                for pos, rid in enumerate(ids, start=1):
                    cur.execute("UPDATE movimientos SET order_index=%s WHERE id_movimiento=%s", (pos, rid))
            conn.commit()
        return {"ok": True}

    @app.post("/estado-cuenta/mover/<cliente>/<registro_id>/<direccion>")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def mover_registro(cliente: str, registro_id: str, direccion: str):
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    (
                        "SELECT id_movimiento, COALESCE(order_index,0) AS ord FROM movimientos m "
                        "JOIN clientes c ON c.id_cliente=m.id_cliente "
                        "WHERE c.nombre_completo=%s AND m.anulada_en IS NULL "
                        "ORDER BY ord, m.fecha, m.id_movimiento"
                    ),
                    (cliente,),
                )
                rows = cur.fetchall()
                ids = [str(r["id_movimiento"]) for r in rows]

                if not ids:
                    flash("Cliente sin registros.", "error")
                    return redirect(url_for("estado.ver_tabla_estado_cuenta", cliente=cliente))

                pos_map = {rid: i for i, rid in enumerate(ids)}
                if registro_id not in pos_map:
                    flash("Registro no encontrado.", "error")
                    return redirect(url_for("estado.ver_tabla_estado_cuenta", cliente=cliente))

                idx = pos_map[registro_id]
                
                if direccion == "up" and idx > 0:
                    ids[idx - 1], ids[idx] = ids[idx], ids[idx - 1]
                elif direccion == "down" and idx < len(ids) - 1:
                    ids[idx + 1], ids[idx] = ids[idx], ids[idx + 1]
                else:
                    # No hay cambios válidos, redirigir sin hacer nada
                    return redirect(url_for("estado.ver_tabla_estado_cuenta", cliente=cliente))

                # Persistir nuevo orden secuencial
                for pos, rid in enumerate(ids, start=1):
                    cur.execute("UPDATE movimientos SET order_index=%s WHERE id_movimiento=%s", (pos, rid))
                
                conn.commit()

        flash("Orden actualizado.", "success")
        return redirect(url_for("estado.ver_tabla_estado_cuenta", cliente=cliente))

    @app.post("/estado-cuenta/eliminar/<registro_id>")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def eliminar_registro(registro_id: str):
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM movimientos WHERE id_movimiento=%s", (registro_id,))
                conn.commit()
        flash("Registro eliminado.", "success")
        return redirect(url_for("estado.ver_tabla_estado_cuenta"))

    @app.post("/estado-cuenta/marcar/<registro_id>/<col>")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def marcar_celda(registro_id: str, col: str):
        col = col.upper()
        if col not in {"DEBE", "HABER", "SALDO"}:
            return {"ok": False, "error": "Columna inválida"}, 400
        payload = request.get_json(silent=True) or {}
        # mark: 0 limpia, 1..9 aplica color índice
        try:
            mark_value = int(payload.get("mark", -1))
        except Exception:
            mark_value = -1
        field = {"DEBE": "mark_debe", "HABER": "mark_haber", "SALDO": "mark_saldo"}[col]
        with db_connect() as conn:
            with conn.cursor() as cur:
                if mark_value == -1:
                    cur.execute(f"SELECT COALESCE({field},0) FROM movimientos WHERE id_movimiento=%s", (registro_id,))
                    current = int(cur.fetchone()[0]) if cur.rowcount else 0
                    newv = 0 if current else 1
                    cur.execute(f"UPDATE movimientos SET {field}=%s WHERE id_movimiento=%s", (newv, registro_id))
                    conn.commit()
                    return {"ok": True, "mark": newv}
                else:
                    newv = max(0, min(9, int(mark_value)))
                    cur.execute(f"UPDATE movimientos SET {field}=%s WHERE id_movimiento=%s", (newv, registro_id))
                    conn.commit()
                    return {"ok": True, "mark": newv}

    @app.post("/estado-cuenta/pagado/<registro_id>")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def marcar_pagado(registro_id: str):
        # Marca la fila como pagada hoy
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE movimientos SET pagado_en=NOW()::date WHERE id_movimiento=%s", (registro_id,))
                conn.commit()
        return {"ok": True}

    @app.post("/estado-cuenta/limpiar-marcados/<cliente>")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def limpiar_marcados(cliente: str):
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE movimientos SET mark_debe=0, mark_haber=0, mark_saldo=0 FROM clientes c WHERE c.id_cliente=movimientos.id_cliente AND c.nombre_completo=%s", (cliente,))
                conn.commit()
            flash("Colores limpiados para el cliente.", "success")
        return redirect(url_for("estado.ver_tabla_estado_cuenta", cliente=cliente))

    @app.post("/estado-cuenta/limpiar-marcados-todos")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def limpiar_marcados_todos():
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE movimientos SET mark_debe=0, mark_haber=0, mark_saldo=0")
                conn.commit()
            flash("Se limpiaron los colores de todos los clientes.", "success")
        return redirect(url_for("estado.ver_tabla_estado_cuenta"))

    @app.post("/estado-cuenta/anular/<registro_id>")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def anular_registro(registro_id: str):
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE movimientos SET anulada_en=NOW() WHERE id_movimiento=%s", (registro_id,))
                conn.commit()
        flash("Registro anulado y movido a 'Anuladas'.", "success")
        return redirect(url_for("estado.ver_tabla_estado_cuenta"))

    # La vista de anuladas se sirve desde el blueprint `estado`

    @app.get("/estado-cuenta/exportar")
    @login_required
    @permission_required(PERM_ESTADO_VIEW)
    def exportar_estado_cuenta():
        # Además de ver, requerir permiso explícito de exportación si se quiere restringir
        if not usuario_tiene_permiso(usuario_actual(), PERM_ESTADO_EXPORT):
            flash("No tiene permisos para exportar.", "error")
            return redirect(url_for("estado.ver_tabla_estado_cuenta"))
        cliente_q = request.args.get("cliente", "").strip()
        if not cliente_q:
            flash("Debe seleccionar un cliente para exportar.", "error")
            return redirect(url_for("estado.ver_tabla_estado_cuenta"))
        start_q = request.args.get("inicio", "").strip()
        end_q = request.args.get("fin", "").strip()
        sort_q = request.args.get("orden", "asc").strip().lower()

        # Reutilizar lógica de ver_tabla_estado_cuenta para obtener df final del cliente
        registros = cargar_json_estado_de_cuenta()
        if not registros:
            flash("No hay datos para exportar.", "info")
            return redirect(url_for("estado.ver_tabla_estado_cuenta"))

        df = pd.DataFrame(registros)
        for col in COLUMNAS_SALIDA + ["CLIENTE"]:
            if col not in df.columns:
                df[col] = "" if col not in ["DEBE", "HABER", "SALDO"] else 0.0
        df["DEBE"] = pd.to_numeric(df["DEBE"], errors="coerce").fillna(0.0)
        df["HABER"] = pd.to_numeric(df["HABER"], errors="coerce").fillna(0.0)
        df["SALDO"] = pd.to_numeric(df["SALDO"], errors="coerce").fillna(0.0)
        df["_FECHA_DT"] = pd.to_datetime(df["FECHA"], errors="coerce")

        df = df[df["CLIENTE"].astype(str).str.strip() == cliente_q]
        if start_q:
            start_dt = pd.to_datetime(start_q, errors="coerce")
            if pd.notna(start_dt):
                df = df[df["_FECHA_DT"] >= start_dt]
        if end_q:
            end_dt = pd.to_datetime(end_q, errors="coerce")
            if pd.notna(end_dt):
                df = df[df["_FECHA_DT"] <= end_dt]

        df_calc = df.copy()
        df_calc["_ORDER"] = pd.to_numeric(df_calc.get("ORDER"), errors="coerce")
        df_calc = df_calc.sort_values(by=["CLIENTE", "_ORDER", "_FECHA_DT", "DOCTO"], ascending=[True, True, True, True], na_position="last")
        df_calc["_MOV"] = df_calc["HABER"] - df_calc["DEBE"]
        df_calc["SALDO"] = df_calc.groupby("CLIENTE")["_MOV"].cumsum().round(2)
        df_calc["_INI"] = 1
        saldos_ini = cargar_saldos_iniciales()
        if saldos_ini:
            def agregar_ini_export(grupo: pd.DataFrame) -> pd.DataFrame:
                cliente = str(grupo["CLIENTE"].iloc[0])
                info = saldos_ini.get(cliente)
                if not info:
                    return grupo
                monto = float(info.get("monto", 0) or 0)
                lado = str(info.get("lado", "haber")).lower()
                mov = monto if lado == "haber" else -monto
                fecha_dt = pd.to_datetime(info.get("fecha", "2025-01-01"), errors="coerce")
                if pd.isna(fecha_dt):
                    fecha_dt = pd.to_datetime("2025-01-01")
                fila_ini = {
                    "CLIENTE": cliente,
                    "FECHA": fecha_dt.strftime("%d/%m/%Y"),
                    "_FECHA_DT": fecha_dt,
                    "DOCTO": "SALDO ANTERIOR",
                    "DETALLE": f"{cliente} - SALDO ANTERIOR",
                    "INT. AG.": "",
                    "DIM": "",
                    "DEBE": monto if lado == "debe" else 0.0,
                    "HABER": monto if lado == "haber" else 0.0,
                    "_MOV": mov,
                    "SALDO": 0.0,
                    "_INI": 0,
                }
                nuevo = pd.concat([pd.DataFrame([fila_ini]), grupo], ignore_index=True)
                nuevo["SALDO"] = nuevo["_MOV"].cumsum().round(2)
                return nuevo
            df_calc = df_calc.groupby("CLIENTE", as_index=False, group_keys=False).apply(agregar_ini_export)
        # Orden final: saldo anterior arriba, luego fecha/DOCTO según sort
        asc = sort_q != "desc"
        if "_INI" not in df_calc.columns:
            df_calc["_INI"] = 1
        df_calc = df_calc.sort_values(by=["_INI", "_FECHA_DT", "DOCTO"], ascending=[True, asc, asc], na_position="last")

        df_calc["FECHA"] = df_calc["_FECHA_DT"].dt.strftime("%d/%m/%Y").fillna("")
        columnas_vista = ["FECHA", "DOCTO", "DETALLE", "INT. AG.", "DIM", "DEBE", "HABER", "SALDO"]
        df_view = df_calc[columnas_vista]

        # Crear Excel con formato
        wb = Workbook()
        ws = wb.active
        ws.title = "Estado de Cuenta"

        # Encabezado superior (centrado y fusionado)
        max_col = 8
        def merge_write(row, text, bold=True, size=12):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
            c = ws.cell(row=row, column=1, value=text)
            c.font = Font(bold=bold, size=size)
            c.alignment = Alignment(horizontal="center")

        merge_write(1, "ESTADO DE CUENTA CLIENTE", bold=True, size=14)
        merge_write(2, "ECC-001", bold=False, size=11)
        merge_write(3, cliente_q.upper(), bold=True, size=13)
        # Fecha de corte: usar 'fin' si vino, si no la última fecha del df
        corte_dt = pd.to_datetime(end_q, errors="coerce")
        if pd.isna(corte_dt):
            if not df_calc.empty and pd.notna(df_calc["_FECHA_DT"].max()):
                corte_dt = df_calc["_FECHA_DT"].max()
            else:
                corte_dt = pd.Timestamp.today()
        merge_write(4, f"AL {corte_dt.strftime('%d DE %B DE %Y').upper()}", bold=False, size=11)
        merge_write(5, "(Expresado en Bolivianos)", bold=False, size=10)

        # Encabezado de tabla
        start_data_row = 7
        headers = ["FECHA", "DOCTO.", "DETALLE", "INT. AG.", "DUI", "DEBE", "HABER", "SALDO"]
        for idx, h in enumerate(headers, start=1):
            ws.cell(row=start_data_row, column=idx, value=h)
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="F1F5F9")
        thin = Side(style="thin", color="E5E7EB")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=start_data_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Datos
        num_fmt = "#,##0.00"
        haber_col_idx = headers.index("HABER") + 1
        rptr = start_data_row + 1
        for _, row in df_view.iterrows():
            ws.cell(row=rptr, column=1, value=row.get("FECHA", ""))
            ws.cell(row=rptr, column=2, value=row.get("DOCTO", ""))
            ws.cell(row=rptr, column=3, value=row.get("DETALLE", ""))
            ws.cell(row=rptr, column=4, value=row.get("INT. AG.", ""))
            ws.cell(row=rptr, column=5, value=row.get("DIM", ""))
            ws.cell(row=rptr, column=6, value=float(row.get("DEBE", 0) or 0))
            ws.cell(row=rptr, column=7, value=float(row.get("HABER", 0) or 0))
            ws.cell(row=rptr, column=8, value=float(row.get("SALDO", 0) or 0))
            rptr += 1
        # Formatos numéricos y relleno en HABER
        for r in ws.iter_rows(min_row=start_data_row + 1, max_row=rptr - 1, min_col=6, max_col=8):
            for c in r:
                c.number_format = num_fmt
                c.border = border
                if c.col_idx == haber_col_idx:
                    c.fill = PatternFill("solid", fgColor="FFF59E")  # amarillo suave
        # Bordes para resto de celdas
        for r in ws.iter_rows(min_row=start_data_row + 1, max_row=rptr - 1, min_col=1, max_col=5):
            for c in r:
                c.border = border

        # Anchos
        widths = [12, 12, 50, 10, 18, 14, 14, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        ws.freeze_panes = f"A{start_data_row + 1}"

        # Totales al final
        sum_debe = float(df_view["DEBE"].sum())
        sum_haber = float(df_view["HABER"].sum())
        saldo_final = float(df_view["SALDO"].iloc[-1]) if not df_view.empty else 0.0

        def write_row(label, val_debe=None, val_haber=None, bold=True, border_box=True):
            nonlocal rptr
            ws.cell(row=rptr, column=1, value=label).font = Font(bold=bold)
            if val_debe is not None:
                c = ws.cell(row=rptr, column=6, value=val_debe)
                c.number_format = num_fmt
            if val_haber is not None:
                c2 = ws.cell(row=rptr, column=7, value=val_haber)
                c2.number_format = num_fmt
            # caja alrededor de las dos celdas numéricas
            if border_box:
                thick = Side(style="thin", color="111111")
                b = Border(top=thick, left=thick, right=thick, bottom=thick)
                for col in (6, 7):
                    ws.cell(row=rptr, column=col).border = b
            rptr += 2  # dejar una fila en blanco

        write_row("SUMAS TOTALES", sum_debe, sum_haber)
        # Saldo a favor del cliente o de la empresa
        saldo_label = f"SALDO A FAVOR DE {cliente_q.upper()}  AL  {corte_dt.strftime('%d/%m/%Y')}"
        if saldo_final < 0:
            write_row(saldo_label, abs(saldo_final), 0.0)
        else:
            write_row(saldo_label, 0.0, saldo_final)
        # Sumas totales e iguales
        total_debe_eq = sum_debe + (abs(saldo_final) if saldo_final < 0 else 0.0)
        total_haber_eq = sum_haber + (saldo_final if saldo_final > 0 else 0.0)
        write_row("SUMAS TOTALES E IGUALES", total_debe_eq, total_haber_eq)

        # Responder archivo
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"estado_cuenta_{cliente_q}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ==========================
    # Saldos Anteriores (gestión)
    # ==========================

    def obtener_clientes_existentes() -> List[str]:
        try:
            with db_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT nombre_completo FROM clientes ORDER BY nombre_completo ASC")
                    rows = cur.fetchall()
                    return [str(r["nombre_completo"]) for r in rows]
        except Exception:
            return []

    @app.get("/estado-cuenta/preferencias")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def ver_prefs_clientes():
        # Lista editable de días de mora por cliente
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT nombre_completo, COALESCE(cp.dias_mora,30) AS dias_mora FROM clientes c LEFT JOIN cliente_prefs cp ON cp.id_cliente=c.id_cliente ORDER BY nombre_completo ASC")
                rows = cur.fetchall()
        return render_template("prefs_clientes.html", filas=rows)

    @app.post("/estado-cuenta/preferencias")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def guardar_prefs_clientes():
        # Guarda días de mora por cliente
        nombre = request.form.get("cliente", "").strip()
        dias = request.form.get("dias", "30").strip()
        try:
            dias_i = max(0, int(dias))
        except Exception:
            flash("Días inválidos.", "error")
            return redirect(url_for("ver_prefs_clientes"))
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id_cliente FROM clientes WHERE nombre_completo=%s", (nombre,))
                row = cur.fetchone()
                if not row:
                    flash("Cliente no existe.", "error")
                    return redirect(url_for("ver_prefs_clientes"))
                cid = row["id_cliente"]
                cur.execute(
                    "INSERT INTO cliente_prefs (id_cliente, dias_mora) VALUES (%s,%s) ON CONFLICT (id_cliente) DO UPDATE SET dias_mora=EXCLUDED.dias_mora",
                    (cid, dias_i),
                )
                conn.commit()
        flash("Preferencia guardada.", "success")
        return redirect(url_for("ver_prefs_clientes"))

    @app.get("/estado-cuenta/saldos-anteriores")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def ver_saldos_anteriores():
        data = cargar_saldos_iniciales()  # {cliente: {monto,lado,fecha}}
        clientes = obtener_clientes_existentes()
        # Convertir a lista para la tabla
        filas = []
        for cliente, info in data.items():
            filas.append({
                "CLIENTE": cliente,
                "FECHA": info.get("fecha", "2025-01-01"),
                "LADO": info.get("lado", "haber"),
                "MONTO": round(float(info.get("monto", 0) or 0), 2),
            })
        # Ordenar por cliente
        filas = sorted(filas, key=lambda r: r["CLIENTE"].lower())
        return render_template(
            "saldos_anteriores.html",
            filas=filas,
            clientes=clientes,
        )

    @app.post("/estado-cuenta/saldos-anteriores")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def guardar_saldo_anterior():
        cliente = request.form.get("cliente", "").strip()
        monto = request.form.get("monto", "").strip()
        lado = request.form.get("lado", "haber").strip().lower()
        fecha = request.form.get("fecha", "2025-01-01").strip() or "2025-01-01"
        if not cliente:
            flash("Seleccione un cliente.", "error")
            return redirect(url_for("ver_saldos_anteriores"))
        try:
            valor = float(str(monto).replace(".", "").replace(",", ".")) if monto else 0.0
        except Exception:
            flash("Monto inválido.", "error")
            return redirect(url_for("ver_saldos_anteriores"))
        if lado not in ("debe", "haber"):
            lado = "haber" if valor >= 0 else "debe"
        data = cargar_saldos_iniciales()
        data[cliente] = {"monto": round(abs(valor), 2), "lado": lado, "fecha": fecha}
        guardar_saldos_iniciales(data)
        flash("Saldo anterior guardado.", "success")
        return redirect(url_for("ver_saldos_anteriores"))

    @app.post("/estado-cuenta/saldos-anteriores/eliminar")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def eliminar_saldo_anterior():
        cliente = request.form.get("cliente", "").strip()
        data = cargar_saldos_iniciales()
        if cliente in data:
            del data[cliente]
            guardar_saldos_iniciales(data)
            flash("Saldo anterior eliminado.", "success")
        else:
            flash("Cliente no encontrado en saldos anteriores.", "error")
        return redirect(url_for("ver_saldos_anteriores"))

    @app.get("/estado-cuenta/editar/<registro_id>")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def editar_registro(registro_id: str):
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT m.id_movimiento AS "ID",
                           c.nombre_completo AS "CLIENTE",
                           COALESCE(TO_CHAR(m.fecha,'YYYY-MM-DD'),'') AS "FECHA",
                           m.docto AS "DOCTO",
                           m.descripcion AS "DETALLE",
                           COALESCE(TO_CHAR(m.condicion_de_pago,'YYYY-MM-DD'),'') AS "CONDICION DE PAGO",
                           m.int_ag AS "INT. AG.",
                           m.dim AS "DIM",
                           m.tipo_de_movimiento,
                           COALESCE(m.monto,0) AS monto
                    FROM movimientos m
                    JOIN clientes c ON c.id_cliente=m.id_cliente
                    WHERE m.id_movimiento=%s
                    """,
                    (registro_id,)
                )
                r = cur.fetchone()
                if not r:
                    flash("Registro no encontrado.", "error")
                    return redirect(url_for("estado.ver_tabla_estado_cuenta"))
                debe = float(r["monto"]) if r["tipo_de_movimiento"] == "CARGO" else 0.0
                haber = float(r["monto"]) if r["tipo_de_movimiento"] == "PAGO" else 0.0
                reg = {
                    "ID": str(r["ID"]),
                    "CLIENTE": r["CLIENTE"],
                    "FECHA": r["FECHA"],
                    "DOCTO": r["DOCTO"] or "",
                    "DETALLE": r["DETALLE"] or "",
                    "CONDICION DE PAGO": r["CONDICION DE PAGO"] or "",
                    "INT. AG.": r["INT. AG."] or "",
                    "DIM": r["DIM"] or "",
                    "DEBE": round(debe, 2),
                    "HABER": round(haber, 2),
                }
        return render_template("editar_registro.html", registro=reg)

    @app.post("/estado-cuenta/editar/<registro_id>")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def editar_registro_post(registro_id: str):
        cliente = request.form.get("CLIENTE", "").strip()
        fecha = request.form.get("FECHA", "").strip() or None
        docto = request.form.get("DOCTO", "").strip() or None
        detalle = request.form.get("DETALLE", "").strip() or None
        debe = request.form.get("DEBE", "").strip()
        haber = request.form.get("HABER", "").strip()
        int_ag = request.form.get("INT. AG.", "").strip() or None
        dim = request.form.get("DIM", "").strip() or None
        venc = request.form.get("CONDICION DE PAGO", "").strip() or None

        # normalizar números
        def to_float(txt: str) -> float:
            if not txt:
                return 0.0
            return float(str(txt).replace(".", "").replace(",", "."))

        v_debe = round(to_float(debe), 2)
        v_haber = round(to_float(haber), 2)
        tipo = "CARGO" if (v_debe > 0 and v_haber == 0) else ("PAGO" if (v_haber > 0 and v_debe == 0) else ("PAGO" if (v_haber - v_debe) >= 0 else "CARGO"))
        monto = round(abs(v_haber - v_debe), 2) if (v_debe > 0 and v_haber > 0) else (v_debe if tipo == "CARGO" else v_haber)

        with db_connect() as conn:
            with conn.cursor() as cur:
                # actualizar cliente si cambió (opcional)
                if cliente:
                    cur.execute("SELECT id_cliente FROM clientes WHERE nombre_completo=%s", (cliente,))
                    row = cur.fetchone()
                    if not row:
                        cur.execute("INSERT INTO clientes (nombre_completo) VALUES (%s) RETURNING id_cliente", (cliente,))
                        idc = cur.fetchone()["id_cliente"]
                    else:
                        idc = row["id_cliente"]
                    cur.execute(
                        """
                        UPDATE movimientos SET id_cliente=%s, fecha=%s, tipo_de_movimiento=%s, monto=%s,
                               descripcion=%s, docto=%s, int_ag=%s, dim=%s, condicion_de_pago=%s
                        WHERE id_movimiento=%s
                        """,
                        (idc, fecha, tipo, monto, detalle, docto, int_ag, dim, venc, registro_id)
                    )
                else:
                    cur.execute(
                        """
                        UPDATE movimientos SET fecha=%s, tipo_de_movimiento=%s, monto=%s,
                               descripcion=%s, docto=%s, int_ag=%s, dim=%s, condicion_de_pago=%s
                        WHERE id_movimiento=%s
                        """,
                        (fecha, tipo, monto, detalle, docto, int_ag, dim, venc, registro_id)
                    )
                conn.commit()
        flash("Registro actualizado.", "success")
        return redirect(url_for("estado.ver_tabla_estado_cuenta"))

    @app.post("/estado-cuenta/vaciar")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def vaciar_estado_cuenta():
        try:
            with db_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute("TRUNCATE TABLE movimientos RESTART IDENTITY")
                    conn.commit()
            flash("Todos los registros de Estado de Cuenta (DB) fueron eliminados.", "success")
        except Exception as exc:
            flash(f"No se pudo vaciar: {exc}", "error")
        return redirect(url_for("estado.ver_tabla_estado_cuenta"))

    @app.post("/estado-cuenta/saldo-inicial")
    @login_required
    @permission_required(PERM_ESTADO_UPLOAD)
    def set_saldo_inicial():
        cliente = request.form.get("cliente", "").strip()
        monto = request.form.get("monto", "").strip()
        lado = request.form.get("lado", "haber").strip().lower()
        if not cliente:
            flash("Debe seleccionar un cliente.", "error")
            return redirect(url_for("estado.ver_tabla_estado_cuenta"))
        try:
            valor = float(monto.replace(".", "").replace(",", ".")) if monto else 0.0
        except Exception:
            flash("Monto inválido.", "error")
            return redirect(url_for("estado.ver_tabla_estado_cuenta", cliente=cliente))
        if lado not in ("debe", "haber"):
            lado = "haber" if valor >= 0 else "debe"
        cid = db_get_or_create_cliente(cliente)
        if cid:
            with db_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO saldos_iniciales (id_cliente, monto, lado, fecha) 
                        VALUES (%s,%s,%s,%s)
                        ON CONFLICT (id_cliente) DO UPDATE SET monto=EXCLUDED.monto,lado=EXCLUDED.lado,fecha=EXCLUDED.fecha
                        """,
                        (cid, round(abs(valor), 2), lado, "2025-01-01"),
                    )
                    conn.commit()
        flash("Saldo inicial actualizado.", "success")
        return redirect(url_for("estado.ver_tabla_estado_cuenta", cliente=cliente))

    @app.get("/planilla")
    @login_required
    @permission_required(PERM_PLANILLA_VIEW)
    def modulo_planilla():
        return render_template("planilla_menu.html")

    @app.get("/planilla/empleados")
    @login_required
    @permission_required(PERM_PLANILLA_EMPL_VIEW)
    def planilla_empleados():
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM empleados ORDER BY nombres_apellidos ASC")
                rows = cur.fetchall()
        return render_template("planilla_empleados.html", empleados=rows)

    @app.post("/planilla/empleados")
    @login_required
    @permission_required(PERM_PLANILLA_EMPL_VIEW)
    def planilla_empleados_post():
        f = request.form
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO empleados (carnet, nombres_apellidos, cargo, cua, cns, cns_patronal, fecha_ingreso, haber_basico)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (carnet) DO UPDATE SET
                        nombres_apellidos=EXCLUDED.nombres_apellidos,
                        cargo=EXCLUDED.cargo,
                        cua=EXCLUDED.cua,
                        cns=EXCLUDED.cns,
                        cns_patronal=EXCLUDED.cns_patronal,
                        fecha_ingreso=EXCLUDED.fecha_ingreso,
                        haber_basico=EXCLUDED.haber_basico
                    RETURNING id_empleado, carnet, nombres_apellidos, cargo, cua, cns, cns_patronal, COALESCE(TO_CHAR(fecha_ingreso,'YYYY-MM-DD'),'') AS fecha_ingreso, haber_basico
                    """,
                    (
                        f.get("carnet"), f.get("nombres"), f.get("cargo"), f.get("cua"), f.get("cns"), f.get("cns_patronal"),
                        f.get("fecha_ingreso"), float((f.get("haber_basico") or 0) or 0)
                    ),
                )
                row = cur.fetchone()
                conn.commit()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return {"ok": True, "empleado": row}, 200
        flash("Empleado guardado.", "success")
        return redirect(url_for("planilla_empleados"))

    @app.get("/planilla/periodo")
    @login_required
    @permission_required(PERM_PLANILLA_PERIODO_VIEW)
    def planilla_periodo():
        periodo = request.args.get("periodo", "")
        # generar lista de meses de la gestión actual (12 meses)
        current = datetime.now()
        year = current.year
        months: List[str] = [f"{year}{m:02d}" for m in range(1, 13)]
        if not periodo:
            periodo = f"{year}{current.month:02d}"
        rows = []
        totals = {"haber_basico":0.0, "bono_antiguedad":0.0, "otros_ingresos":0.0, "total_ganado":0.0,
                  "afp_1271":0.0, "ap_solidario":0.0, "quincena":0.0, "anticipos":0.0, "prestamos":0.0,
                  "entel":0.0, "otros_desc":0.0, "atrasos":0.0, "rc_iva":0.0, "total_afps":0.0,
                  "total_anticipos":0.0, "total_desc":0.0, "total_afp_ant_desc":0.0, "liquido_pagable":0.0}
        with db_connect() as conn:
            with conn.cursor() as cur:
                # LEFT JOIN para mostrar todos los empleados, aunque no tengan item del periodo
                if periodo:
                    cur.execute(
                        """
                        SELECT 
                          e.id_empleado, e.carnet, e.nombres_apellidos, e.cargo, e.cua, e.cns, e.cns_patronal,
                          COALESCE(TO_CHAR(e.fecha_ingreso,'YYYY-MM-DD'),'') AS fecha_ingreso, e.haber_basico,
                          COALESCE(pi.dias_trab,30) AS dias_trab,
                          COALESCE(pi.bono_antiguedad,0) AS bono_antiguedad,
                          COALESCE(pi.otros_ingresos,0) AS otros_ingresos,
                          (e.haber_basico + COALESCE(pi.bono_antiguedad,0) + COALESCE(pi.otros_ingresos,0)) AS total_ganado,
                          ROUND((e.haber_basico + COALESCE(pi.bono_antiguedad,0) + COALESCE(pi.otros_ingresos,0)) * 0.1271, 2) AS afp_1271,
                          COALESCE(pi.ap_solidario,0) AS ap_solidario,
                          COALESCE(pi.quincena,0) AS quincena,
                          COALESCE(pi.anticipos,0) AS anticipos,
                          COALESCE(pi.prestamos,0) AS prestamos,
                          COALESCE(pi.entel,0) AS entel,
                          COALESCE(pi.otros_desc,0) AS otros_desc,
                          COALESCE(pi.atrasos,0) AS atrasos,
                          COALESCE(pi.rc_iva,0) AS rc_iva,
                          (ROUND((e.haber_basico + COALESCE(pi.bono_antiguedad,0) + COALESCE(pi.otros_ingresos,0)) * 0.1271, 2) + COALESCE(pi.ap_solidario,0)) AS total_afps,
                          (COALESCE(pi.quincena,0) + COALESCE(pi.anticipos,0) + COALESCE(pi.prestamos,0)) AS total_anticipos,
                          (COALESCE(pi.entel,0) + COALESCE(pi.otros_desc,0) + COALESCE(pi.atrasos,0) + COALESCE(pi.rc_iva,0)) AS total_desc,
                          ((ROUND((e.haber_basico + COALESCE(pi.bono_antiguedad,0) + COALESCE(pi.otros_ingresos,0)) * 0.1271, 2) + COALESCE(pi.ap_solidario,0)) +
                           (COALESCE(pi.quincena,0) + COALESCE(pi.anticipos,0) + COALESCE(pi.prestamos,0)) +
                           (COALESCE(pi.entel,0) + COALESCE(pi.otros_desc,0) + COALESCE(pi.atrasos,0) + COALESCE(pi.rc_iva,0))) AS total_afp_ant_desc,
                          ((e.haber_basico + COALESCE(pi.bono_antiguedad,0) + COALESCE(pi.otros_ingresos,0)) -
                           ((ROUND((e.haber_basico + COALESCE(pi.bono_antiguedad,0) + COALESCE(pi.otros_ingresos,0)) * 0.1271, 2) + COALESCE(pi.ap_solidario,0)) +
                            (COALESCE(pi.quincena,0) + COALESCE(pi.anticipos,0) + COALESCE(pi.prestamos,0)) +
                            (COALESCE(pi.entel,0) + COALESCE(pi.otros_desc,0) + COALESCE(pi.atrasos,0) + COALESCE(pi.rc_iva,0)))) AS liquido_pagable,
                          COALESCE(pi.rc_iva_acum,0) AS rc_iva_acum
                        FROM empleados e
                        LEFT JOIN planilla_items pi ON pi.id_empleado=e.id_empleado AND pi.periodo_yyyymm=%s
                        ORDER BY e.nombres_apellidos ASC
                        """,
                        (periodo,),
                    )
                    rows = cur.fetchall()
                    # acumular totales
                    for r in rows:
                        get = r.get if isinstance(r, dict) else r.__getitem__
                        totals["haber_basico"] += float(get("haber_basico") or 0)
                        totals["bono_antiguedad"] += float(get("bono_antiguedad") or 0)
                        totals["otros_ingresos"] += float(get("otros_ingresos") or 0)
                        totals["total_ganado"] += float(get("total_ganado") or 0)
                        totals["afp_1271"] += float(get("afp_1271") or 0)
                        totals["ap_solidario"] += float(get("ap_solidario") or 0)
                        totals["quincena"] += float(get("quincena") or 0)
                        totals["anticipos"] += float(get("anticipos") or 0)
                        totals["prestamos"] += float(get("prestamos") or 0)
                        totals["entel"] += float(get("entel") or 0)
                        totals["otros_desc"] += float(get("otros_desc") or 0)
                        totals["atrasos"] += float(get("atrasos") or 0)
                        totals["rc_iva"] += float(get("rc_iva") or 0)
                        totals["total_afps"] += float(get("total_afps") or 0)
                        totals["total_anticipos"] += float(get("total_anticipos") or 0)
                        totals["total_desc"] += float(get("total_desc") or 0)
                        totals["total_afp_ant_desc"] += float(get("total_afp_ant_desc") or 0)
                        totals["liquido_pagable"] += float(get("liquido_pagable") or 0)
                else:
                    cur.execute(
                        "SELECT id_empleado, carnet, nombres_apellidos, cargo, cua, cns, cns_patronal, COALESCE(TO_CHAR(fecha_ingreso,'YYYY-MM-DD'),'') AS fecha_ingreso, haber_basico FROM empleados ORDER BY nombres_apellidos ASC"
                    )
                    rows = cur.fetchall()
        return render_template("planilla_periodo.html", rows=rows, periodo=periodo, totals=totals, months=months)

    @app.post("/planilla/periodo")
    @login_required
    @permission_required(PERM_PLANILLA_PERIODO_VIEW)
    def planilla_periodo_post():
        if not usuario_tiene_permiso(usuario_actual(), PERM_PLANILLA_EDIT):
            flash("No tiene permisos para editar planilla.", "error")
            return redirect(url_for("planilla_periodo", periodo=request.form.get("periodo","")))
        periodo = request.form.get("periodo", "")
        id_empleado = request.form.get("id_empleado")
        dias_trab = int(request.form.get("dias_trab", 30) or 30)
        bono_ant = float(request.form.get("bono_antiguedad", 0) or 0)
        otros_ing = float(request.form.get("otros_ingresos", 0) or 0)
        quincena = float(request.form.get("quincena", 0) or 0)
        anticipos = float(request.form.get("anticipos", 0) or 0)
        prestamos = float(request.form.get("prestamos", 0) or 0)
        entel = float(request.form.get("entel", 0) or 0)
        otros_desc = float(request.form.get("otros_desc", 0) or 0)
        atrasos = float(request.form.get("atrasos", 0) or 0)
        rc_iva = float(request.form.get("rc_iva", 0) or 0)
        with db_connect() as conn:
            with conn.cursor() as cur:
                # obtener haber_basico del empleado
                cur.execute("SELECT haber_basico FROM empleados WHERE id_empleado=%s", (id_empleado,))
                hb = float(cur.fetchone()[0] or 0)
                total_ganado = hb + bono_ant + otros_ing
                afp_1271 = round(total_ganado * 0.1271, 2)
                total_afps = afp_1271  # + ap_solidario (0 por ahora)
                total_anticipos = quincena + anticipos + prestamos
                total_desc = total_anticipos + entel + otros_desc + atrasos + rc_iva + total_afps
                total_afp_ant_desc = total_desc
                liquido = max(0.0, total_ganado - total_desc)
                cur.execute(
                    """
                    INSERT INTO planilla_items (periodo_yyyymm, id_empleado, dias_trab, bono_antiguedad, otros_ingresos, total_ganado,
                                                afp_1271, ap_solidario, quincena, anticipos, prestamos, entel, otros_desc, atrasos,
                                                rc_iva, total_afps, total_anticipos, total_desc, total_afp_ant_desc, liquido_pagable, rc_iva_acum)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0)
                    ON CONFLICT (periodo_yyyymm, id_empleado) DO UPDATE SET
                        dias_trab=EXCLUDED.dias_trab,
                        bono_antiguedad=EXCLUDED.bono_antiguedad,
                        otros_ingresos=EXCLUDED.otros_ingresos,
                        total_ganado=EXCLUDED.total_ganado,
                        afp_1271=EXCLUDED.afp_1271,
                        ap_solidario=EXCLUDED.ap_solidario,
                        quincena=EXCLUDED.quincena,
                        anticipos=EXCLUDED.anticipos,
                        prestamos=EXCLUDED.prestamos,
                        entel=EXCLUDED.entel,
                        otros_desc=EXCLUDED.otros_desc,
                        atrasos=EXCLUDED.atrasos,
                        rc_iva=EXCLUDED.rc_iva,
                        total_afps=EXCLUDED.total_afps,
                        total_anticipos=EXCLUDED.total_anticipos,
                        total_desc=EXCLUDED.total_desc,
                        total_afp_ant_desc=EXCLUDED.total_afp_ant_desc,
                        liquido_pagable=EXCLUDED.liquido_pagable
                    """,
                    (periodo, id_empleado, dias_trab, bono_ant, otros_ing, total_ganado, afp_1271, 0.0, quincena, anticipos, prestamos, entel, otros_desc, atrasos, rc_iva, total_afps, total_anticipos, total_desc, total_afp_ant_desc, liquido),
                )
                conn.commit()
        flash("Planilla guardada.", "success")
        return redirect(url_for("planilla_periodo", periodo=periodo))

    @app.post("/planilla/periodo/snapshot")
    @login_required
    @permission_required(PERM_PLANILLA_VIEW)
    def planilla_snapshot():
        periodo = request.form.get("periodo", "").strip()
        if not periodo:
            return {"ok": False, "error": "Periodo requerido"}, 400
        # construir snapshot del estado actual de la tabla (vista consolidada)
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT row_to_json(t) FROM (SELECT * FROM planilla_items WHERE periodo_yyyymm=%s) t",
                    (periodo,),
                )
                data_rows = [r[0] for r in cur.fetchall()]
                cur.execute(
                    "INSERT INTO planilla_snapshots (periodo_yyyymm, data) VALUES (%s,%s)",
                    (periodo, json.dumps(data_rows)),
                )
                conn.commit()
        # Si viene por AJAX, responder JSON. Si no, redirigir con flash.
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return {"ok": True}
        flash(f"Planilla del periodo {periodo} guardada.", "success")
        return redirect(url_for("planilla_periodo", periodo=periodo))

    @app.post("/planilla/periodo/set")
    @login_required
    @permission_required(PERM_PLANILLA_PERIODO_VIEW)
    def planilla_periodo_set():
        if not usuario_tiene_permiso(usuario_actual(), PERM_PLANILLA_EDIT):
            return {"ok": False, "error": "Sin permisos para editar"}, 403
        try:
            data = request.get_json(silent=True) or {}
            periodo = (data.get("periodo") or "").strip()
            id_empleado = data.get("id_empleado")
            field = (data.get("field") or "").strip()
            raw_value = data.get("value")
            if not periodo or not id_empleado or not field:
                return {"ok": False, "error": "Datos incompletos"}, 400
            allowed_fields = {
                "dias_trab": "int",
                "bono_antiguedad": "num",
                "otros_ingresos": "num",
                "ap_solidario": "num",
                "quincena": "num",
                "anticipos": "num",
                "prestamos": "num",
                "entel": "num",
                "otros_desc": "num",
                "atrasos": "num",
                "rc_iva": "num",
            }
            if field not in allowed_fields:
                return {"ok": False, "error": "Campo no editable"}, 400
            def to_num(x):
                try:
                    return float(str(x).replace(",", ".").replace(" ", ""))
                except Exception:
                    return 0.0
            def to_int(x):
                try:
                    return int(x)
                except Exception:
                    return 0
            value = to_int(raw_value) if allowed_fields[field] == "int" else to_num(raw_value)
            with db_connect() as conn:
                with conn.cursor() as cur:
                    # asegurar item
                    cur.execute(
                        "INSERT INTO planilla_items (periodo_yyyymm, id_empleado) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                        (periodo, id_empleado),
                    )
                    # actualizar campo editable
                    cur.execute(f"UPDATE planilla_items SET {field}=%s WHERE periodo_yyyymm=%s AND id_empleado=%s", (value, periodo, id_empleado))
                    # obtener valores para cálculo
                    cur.execute("SELECT haber_basico FROM empleados WHERE id_empleado=%s", (id_empleado,))
                    row_hb = cur.fetchone()
                    hb = float(((row_hb.get("haber_basico") if isinstance(row_hb, dict) else (row_hb[0] if row_hb else 0)) or 0))
                    cur.execute(
                        """
                        SELECT COALESCE(bono_antiguedad,0), COALESCE(otros_ingresos,0), COALESCE(ap_solidario,0),
                               COALESCE(quincena,0), COALESCE(anticipos,0), COALESCE(prestamos,0), COALESCE(entel,0),
                               COALESCE(otros_desc,0), COALESCE(atrasos,0), COALESCE(rc_iva,0)
                        FROM planilla_items WHERE periodo_yyyymm=%s AND id_empleado=%s
                        """,
                        (periodo, id_empleado),
                    )
                    row = cur.fetchone() or {}
                    if not isinstance(row, dict):
                        # fallback si la conexión no devolviera dict_row
                        row = {
                            "bono_antiguedad": row[0] if len(row) > 0 else 0,
                            "otros_ingresos": row[1] if len(row) > 1 else 0,
                            "ap_solidario": row[2] if len(row) > 2 else 0,
                            "quincena": row[3] if len(row) > 3 else 0,
                            "anticipos": row[4] if len(row) > 4 else 0,
                            "prestamos": row[5] if len(row) > 5 else 0,
                            "entel": row[6] if len(row) > 6 else 0,
                            "otros_desc": row[7] if len(row) > 7 else 0,
                            "atrasos": row[8] if len(row) > 8 else 0,
                            "rc_iva": row[9] if len(row) > 9 else 0,
                        }
                    bono = float(row.get("bono_antiguedad", 0) or 0)
                    otros_ing = float(row.get("otros_ingresos", 0) or 0)
                    ap_soli = float(row.get("ap_solidario", 0) or 0)
                    quinc = float(row.get("quincena", 0) or 0)
                    antic = float(row.get("anticipos", 0) or 0)
                    prest = float(row.get("prestamos", 0) or 0)
                    entel = float(row.get("entel", 0) or 0)
                    otros_desc = float(row.get("otros_desc", 0) or 0)
                    atrasos = float(row.get("atrasos", 0) or 0)
                    rc_iva = float(row.get("rc_iva", 0) or 0)
                    # cálculos
                    total_ganado = round(hb + bono + otros_ing, 2)
                    afp_1271 = round(total_ganado * 0.1271, 2)
                    total_afps = round(afp_1271 + ap_soli, 2)
                    total_anticipos = round(quinc + antic + prest, 2)
                    total_desc = round(entel + otros_desc + atrasos + rc_iva, 2)
                    total_afp_ant_desc = round(total_afps + total_anticipos + total_desc, 2)
                    liquido = round(total_ganado - total_afp_ant_desc, 2)
                    cur.execute(
                        """
                        UPDATE planilla_items SET total_ganado=%s, afp_1271=%s, total_afps=%s, total_anticipos=%s,
                               total_desc=%s, total_afp_ant_desc=%s, liquido_pagable=%s
                        WHERE periodo_yyyymm=%s AND id_empleado=%s
                        """,
                        (total_ganado, afp_1271, total_afps, total_anticipos, total_desc, total_afp_ant_desc, liquido, periodo, id_empleado),
                    )
                    conn.commit()
            return {
                "ok": True,
                "field": field,
                "value": value,
                "calc": {
                    "total_ganado": total_ganado,
                    "afp_1271": afp_1271,
                    "total_afps": total_afps,
                    "total_anticipos": total_anticipos,
                    "total_desc": total_desc,
                    "total_afp_ant_desc": total_afp_ant_desc,
                    "liquido_pagable": liquido,
                }
            }
        except Exception as exc:
            try:
                print("[PLANILLA_SET_ERROR]", repr(exc))
            except Exception:
                pass
            return {"ok": False, "error": str(exc)}, 500

    @app.get("/comisiones")
    @login_required
    @permission_required(PERM_COMISIONES_VIEW)
    def modulo_comisiones():
        flash("[+] El módulo 'Comisiones' se implementará próximamente.", "info")
        return redirect(url_for("index"))

    @app.get("/dim")
    @login_required
    @permission_required(PERM_DIM_VIEW)
    def modulo_dim():
        flash("[+] El módulo 'Datos DIM' se implementará próximamente.", "info")
        return redirect(url_for("index"))

    # ==========================
    # Dashboard (saldos, mora, planilla)
    # ==========================

    @app.get("/dashboard")
    @login_required
    def dashboard():
        # Chequeo de permiso de vista dashboard
        if not usuario_tiene_permiso(usuario_actual(), "dashboard:view"):
            flash("No tiene permisos para el Dashboard.", "error")
            return redirect(url_for("index"))
        # Saldos por cliente
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT c.nombre_completo AS cliente,
                           SUM(CASE WHEN m.tipo_de_movimiento='PAGO' THEN m.monto ELSE 0 END) AS haber,
                           SUM(CASE WHEN m.tipo_de_movimiento='CARGO' THEN m.monto ELSE 0 END) AS debe,
                           SUM(CASE WHEN m.tipo_de_movimiento='PAGO' THEN m.monto ELSE 0 END) -
                           SUM(CASE WHEN m.tipo_de_movimiento='CARGO' THEN m.monto ELSE 0 END) AS saldo
                    FROM movimientos m
                    JOIN clientes c ON c.id_cliente=m.id_cliente
                    WHERE m.anulada_en IS NULL
                    GROUP BY c.nombre_completo
                    """
                )
                saldos_rows = cur.fetchall()

                # Mora por cliente (FA/PG vencidos 30 días, no pagado)
                cur.execute(
                    """
                    SELECT c.nombre_completo AS cliente,
                           COUNT(*) AS mora_count
                    FROM movimientos m
                    JOIN clientes c ON c.id_cliente=m.id_cliente
                    WHERE m.anulada_en IS NULL
                      AND (m.docto ILIKE 'FA%%' OR m.docto ILIKE 'PG%%')
                      AND m.fecha IS NOT NULL
                      AND (m.pagado_en IS NULL)
                      AND (m.fecha + INTERVAL '30 day') < NOW()
                    GROUP BY c.nombre_completo
                    """
                )
                mora_counts_rows = cur.fetchall()

                # Totales de planilla por mes (año actual)
                year = datetime.now().year
                cur.execute(
                    """
                    SELECT periodo_yyyymm,
                           COALESCE(SUM(liquido_pagable),0) AS total_liquido
                    FROM planilla_items
                    WHERE LEFT(periodo_yyyymm,4)=%s
                    GROUP BY periodo_yyyymm
                    """,
                    (str(year),),
                )
                planilla_rows = cur.fetchall()

        # Preparar datos para charts y KPIs
        def getval(row, key):
            return row.get(key) if isinstance(row, dict) else row[0]  # unused here

        # ordenar saldos desc
        saldos_sorted = sorted(
            saldos_rows,
            key=lambda r: (r.get("saldo") if isinstance(r, dict) else r[3]),
            reverse=True,
        )
        chart_clients = [r.get("cliente") if isinstance(r, dict) else r[0] for r in saldos_sorted]
        chart_saldos = [float(r.get("saldo") if isinstance(r, dict) else r[3]) for r in saldos_sorted]

        # KPIs
        total_clientes = len(saldos_rows)
        total_saldo = sum(chart_saldos)
        total_mora_docs = sum(int(r.get("mora_count") if isinstance(r, dict) else r[1]) for r in mora_counts_rows) if mora_counts_rows else 0
        clientes_con_mora = len(mora_counts_rows)

        mora_counts = {
            (r.get("cliente") if isinstance(r, dict) else r[0]): int(r.get("mora_count") if isinstance(r, dict) else r[1])
            for r in mora_counts_rows
        }
        # Top 10 clientes por documentos en mora
        mora_sorted = sorted(mora_counts.items(), key=lambda kv: kv[1], reverse=True)
        mora_top_clients = [kv[0] for kv in mora_sorted[:10]]
        mora_top_counts = [kv[1] for kv in mora_sorted[:10]]

        # meses año actual
        months_labels = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        months_keys = [f"{datetime.now().year}{m:02d}" for m in range(1,13)]
        totals_by_month_map = {(r.get("periodo_yyyymm") if isinstance(r, dict) else r[0]): float(r.get("total_liquido") if isinstance(r, dict) else r[1]) for r in planilla_rows}
        payroll_data = [totals_by_month_map.get(k, 0.0) for k in months_keys]

        clientes_list = [r.get("cliente") if isinstance(r, dict) else r[0] for r in saldos_rows]

        return render_template(
            "dashboard.html",
            saldos=saldos_sorted,
            chart_clients=chart_clients,
            chart_saldos=chart_saldos,
            mora_counts=mora_counts,
            clientes=clientes_list,
            total_clientes=total_clientes,
            total_saldo=total_saldo,
            total_mora_docs=total_mora_docs,
            clientes_con_mora=clientes_con_mora,
            mora_top_clients=mora_top_clients,
            mora_top_counts=mora_top_counts,
            months_labels=months_labels,
            payroll_data=payroll_data,
        )

    @app.get("/dashboard/mora_docs")
    @login_required
    def dashboard_mora_docs():
        if not usuario_tiene_permiso(usuario_actual(), "dashboard:view"):
            return {"ok": False, "error": "Sin permisos"}, 403
        cliente = request.args.get("cliente", "").strip()
        if not cliente:
            return {"ok": False, "error": "Cliente requerido"}, 400
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT m.docto, COALESCE(TO_CHAR(m.fecha,'YYYY-MM-DD'),'') AS fecha,
                           COALESCE(TO_CHAR((m.fecha + INTERVAL '30 day'),'YYYY-MM-DD'),'') AS vencimiento,
                           GREATEST(0, (NOW()::date - (m.fecha + INTERVAL '30 day')::date)) AS dias,
                           m.monto
                    FROM movimientos m
                    JOIN clientes c ON c.id_cliente=m.id_cliente
                    WHERE c.nombre_completo=%s
                      AND m.anulada_en IS NULL
                      AND (m.docto ILIKE 'FA%%' OR m.docto ILIKE 'PG%%')
                      AND m.fecha IS NOT NULL
                      AND (m.pagado_en IS NULL)
                      AND (m.fecha + INTERVAL '30 day') < NOW()
                    ORDER BY m.fecha
                    """,
                    (cliente,),
                )
                rows = cur.fetchall()
        out = [
            {
                "docto": r.get("docto") if isinstance(r, dict) else r[0],
                "fecha": r.get("fecha") if isinstance(r, dict) else r[1],
                "vencimiento": r.get("vencimiento") if isinstance(r, dict) else r[2],
                "dias": int(r.get("dias") if isinstance(r, dict) else r[3]),
                "monto": float(r.get("monto") if isinstance(r, dict) else r[4]),
            }
            for r in rows
        ]
        return {"ok": True, "docs": out}
    # ==========================
    # Autenticación: Login / Logout
    # ==========================

    # Rutas de auth se sirven desde blueprint `auth`

    return app


app = create_app()


if __name__ == "__main__":
    # Modo desarrollo directo con python app.py
    app.run(debug=True)


